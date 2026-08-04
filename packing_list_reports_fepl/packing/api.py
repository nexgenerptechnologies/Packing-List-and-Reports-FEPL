import frappe
from frappe import _
from frappe.utils import flt, getdate
from frappe.utils.xlsxutils import make_xlsx

@frappe.whitelist(allow_guest=False)
def download_supplier_rfq(rfq_name):
	if not frappe.has_permission("Request for Quotation", "read"):
		frappe.throw(_("Not permitted"))

	columns = [
		{"label": "RFQ ID", "fieldname": "rfq_id"},
		{"label": "RFQ Date", "fieldname": "rfq_date"},
		{"label": "Customer Name", "fieldname": "customer_name"},
		{"label": "Project", "fieldname": "project"},
		{"label": "SOP Date", "fieldname": "sop_date"},
		{"label": "Item Code", "fieldname": "item_code"},
		{"label": "Item Name", "fieldname": "item_name"},
		{"label": "Description", "fieldname": "description"},
		{"label": "Brand", "fieldname": "brand"},
		{"label": "Monthly Qty", "fieldname": "monthly_qty"},
		{"label": "Customer Description", "fieldname": "customer_description"},
		{"label": "Sales Partner", "fieldname": "sales_partner"},
		{"label": "Stock", "fieldname": "stock_qty"},
		{"label": "Reserved Stock", "fieldname": "reserved_stock"},
		{"label": "Supplier MPN", "fieldname": "supplier_mpn"},
		{"label": "Supplier Description", "fieldname": "supplier_description"},
		{"label": "MAKE", "fieldname": "make"},
		{"label": "SPQ", "fieldname": "spq"},
		{"label": "Lead Time", "fieldname": "lead_time"},
		{"label": "Price Per 1000", "fieldname": "price_per_1000"},
		{"label": "Quote date", "fieldname": "quote_date"},
		{"label": "Remarks", "fieldname": "remarks"}
	]

	q = frappe.get_doc("Request for Quotation", rfq_name)
	
	item_codes = [d.item_code for d in q.items if d.item_code]
	stock_qty_map = {}
	reserved_stock_map = {}
	
	if item_codes:
		bin_data = frappe.db.sql("""
			SELECT item_code, SUM(actual_qty) AS stock_qty
			FROM `tabBin`
			WHERE item_code IN %s
			GROUP BY item_code
		""", (tuple(item_codes),), as_dict=1)
		stock_qty_map = {b.item_code: flt(b.stock_qty) for b in bin_data}

		reserved_data = frappe.db.sql("""
			SELECT item_code, SUM(reserved_qty - IFNULL(delivered_qty, 0)) AS reserved_qty
			FROM `tabStock Reservation Entry`
			WHERE docstatus = 1 AND item_code IN %s
			GROUP BY item_code
		""", (tuple(item_codes),), as_dict=1)
		reserved_stock_map = {r.item_code: flt(r.reserved_qty) for r in reserved_data}

	data = [[c["label"] for c in columns]]
	for item in q.items:
		brand = item.get("brand") or frappe.db.get_value("Item", item.item_code, "brand") or ""
		sop = item.get("custom_sop_date") or q.get("custom_sop_date") or ""
		mq = item.get("custom_monthly_qty") or item.qty or 0.0
		cd = item.get("custom_customer_description") or ""
		
		row_dict = {
			"rfq_id": q.name,
			"rfq_date": str(q.transaction_date) if q.transaction_date else "",
			"customer_name": q.get("custom_customer_name") or "",
			"project": q.get("custom_project") or "",
			"sop_date": str(sop) if sop else "",
			"item_code": item.item_code,
			"item_name": item.item_name or "",
			"description": item.description or "",
			"brand": brand,
			"monthly_qty": mq,
			"customer_description": cd,
			"sales_partner": item.get("custom_sales_partner") or "",
			"stock_qty": stock_qty_map.get(item.item_code, 0.0),
			"reserved_stock": reserved_stock_map.get(item.item_code, 0.0),
			"supplier_mpn": "",
			"supplier_description": "",
			"make": "",
			"spq": "",
			"lead_time": "",
			"price_per_1000": "",
			"quote_date": "",
			"remarks": ""
		}
		
		row_list = [row_dict.get(c["fieldname"]) for c in columns]
		data.append(row_list)

	xlsx_file = make_xlsx(data, "Supplier RFQ")
	
	frappe.response["filename"] = f"{rfq_name}_Supplier_RFQ.xlsx"
	frappe.response["filecontent"] = xlsx_file.getvalue()
	frappe.response["type"] = "binary"

@frappe.whitelist(allow_guest=False)
def upload_supplier_excel(rfq_name, file_url):
	import openpyxl
	from frappe.utils.file_manager import get_file_path
	import datetime

	if not frappe.has_permission("Request for Quotation", "write"):
		frappe.throw(_("Not permitted"))

	file_path = get_file_path(file_url)
	
	try:
		wb = openpyxl.load_workbook(file_path, data_only=True)
		ws = wb.active
	except Exception as e:
		frappe.throw(_("Failed to read Excel file: {0}").format(str(e)))

	rfq = frappe.get_doc("Request for Quotation", rfq_name)
	
	# Find headers
	headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
	
	def get_idx(name):
		return headers.index(name) if name in headers else -1

	ic_idx = get_idx("Item Code")
	mpn_idx = get_idx("Supplier MPN")
	desc_idx = get_idx("Supplier Description")
	make_idx = get_idx("MAKE")
	spq_idx = get_idx("SPQ")
	lt_idx = get_idx("Lead Time")
	price_idx = get_idx("Price Per 1000")
	date_idx = get_idx("Quote date")
	remarks_idx = get_idx("Remarks")

	if ic_idx == -1:
		frappe.throw(_("Invalid Template: Missing 'Item Code' column."))

	sales_partners_to_notify = set()
	updated_items = 0

	# Update Items
	for row in ws.iter_rows(min_row=2, values_only=True):
		item_code = row[ic_idx]
		if not item_code: continue

		for item in rfq.items:
			if item.item_code == item_code:
				if mpn_idx != -1: item.custom_supplier_mpn = str(row[mpn_idx] or "")
				if desc_idx != -1: item.custom_supplier_description = str(row[desc_idx] or "")
				if make_idx != -1: item.custom_make = str(row[make_idx] or "")
				if spq_idx != -1: item.custom_spq = flt(row[spq_idx])
				if lt_idx != -1: item.custom_lead_time = str(row[lt_idx] or "")
				if price_idx != -1: item.custom_price_per_1000 = flt(row[price_idx])
				if remarks_idx != -1: item.custom_remarks = str(row[remarks_idx] or "")
				
				if date_idx != -1:
					q_date = row[date_idx]
					if isinstance(q_date, datetime.datetime):
						item.custom_quote_date = q_date.date()
					elif q_date:
						try:
							item.custom_quote_date = getdate(q_date)
						except Exception:
							pass
						if isinstance(q_date, str):
							q_str = q_date.replace('/', '-')
							parts = q_str.split('-')
							if len(parts) == 3:
								if len(parts[0]) <= 2 and int(parts[0]) > 12:
									item.custom_quote_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
								elif len(parts[0]) <= 2 and len(parts[2]) == 4:
									item.custom_quote_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
								else:
									item.custom_quote_date = q_date
				
				if item.custom_sales_partner:
					sales_partners_to_notify.add(item.custom_sales_partner)
					
				updated_items += 1
				break
				
	if updated_items > 0:
		rfq.save(ignore_permissions=True)
		
		# Notification 2: Notify Sales Partners
		for sp_name in sales_partners_to_notify:
			user_email = ""
			try:
				sp = frappe.get_doc("Sales Partner", sp_name)
				if hasattr(sp, "user") and sp.user:
					user_email = sp.user
				elif hasattr(sp, "partner_email") and sp.partner_email:
					user_email = sp.partner_email
			except Exception:
				pass
				
			if not user_email:
				user_email = frappe.db.get_value("User", {"full_name": sp_name})
			
			if user_email and frappe.db.exists("User", user_email):
				doc = frappe.new_doc("Notification Log")
				doc.subject = _("Supplier Pricing Uploaded for RFQ: {0}").format(rfq.name)
				doc.email_content = _("The Product Manager has uploaded supplier pricing for your requested items on {0}.").format(rfq.name)
				doc.for_user = user_email
				doc.document_type = "Request for Quotation"
				doc.document_name = rfq.name
				doc.insert(ignore_permissions=True)
				frappe.publish_realtime('notification', doc.subject, user=user_email)

	return "Success"

@frappe.whitelist(allow_guest=False)
def upload_quotation_excel(file_url):
	import openpyxl
	from frappe.utils.file_manager import get_file_path
	import datetime

	if not frappe.has_permission("Quotation", "read"):
		frappe.throw(_("Not permitted"))

	file_path = get_file_path(file_url)
	
	try:
		wb = openpyxl.load_workbook(file_path, data_only=True)
		ws = wb.active
	except Exception as e:
		frappe.throw(_("Failed to read Excel file: {0}").format(str(e)))

	# Find headers (case-insensitive & stripped)
	headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
	
	def get_idx(candidates):
		if isinstance(candidates, str):
			candidates = [candidates]
		for c in candidates:
			c_lower = c.lower()
			for idx, h in enumerate(headers):
				if h.lower() == c_lower:
					return idx
		return -1

	cust_idx = get_idx(["Customer Name", "Customer"])
	proj_idx = get_idx(["Project", "Project Name"])
	sop_idx = get_idx(["SOP Date", "SOP"])
	ic_idx = get_idx(["Item Code", "Part Number", "Part No"])
	in_idx = get_idx(["Item Name", "Part Name"])
	desc_idx = get_idx(["Description", "Item Description"])
	brand_idx = get_idx(["Brand", "Make"])
	mq_idx = get_idx(["Monthly Qty", "Monthly Quantity", "Qty", "Quantity"])
	cd_idx = get_idx(["Customer Description", "Cust Description", "Cust Desc"])
	sp_idx = get_idx(["Sales Partner", "Partner"])

	header_customer = ""
	header_project = ""

	items = []
	unmatched = []

	for row in ws.iter_rows(min_row=2, values_only=True):
		# Skip completely empty rows
		if not any(row):
			continue

		row_cust = str(row[cust_idx]).strip() if cust_idx != -1 and row[cust_idx] else ""
		row_proj = str(row[proj_idx]).strip() if proj_idx != -1 and row[proj_idx] else ""
		if row_cust and not header_customer:
			match_cust = frappe.db.get_value("Customer", {"customer_name": row_cust}, "name") or frappe.db.get_value("Customer", {"name": row_cust}, "name")
			header_customer = match_cust or row_cust
		if row_proj and not header_project:
			header_project = row_proj

		raw_ic = str(row[ic_idx]).strip() if ic_idx != -1 and row[ic_idx] else ""
		raw_in = str(row[in_idx]).strip() if in_idx != -1 and row[in_idx] else ""
		raw_desc = str(row[desc_idx]).strip() if desc_idx != -1 and row[desc_idx] else ""

		# 3-Way Item Matching Logic:
		matched_item = None
		# 1. Match on Item Code
		if raw_ic:
			matched_item = frappe.db.get_value("Item", {"name": raw_ic}, ["name", "item_name", "description", "stock_uom", "custom_msp", "brand"], as_dict=1)
			if not matched_item:
				matched_item = frappe.db.get_value("Item", {"item_code": raw_ic}, ["name", "item_code", "item_name", "description", "stock_uom", "custom_msp", "brand"], as_dict=1)

		# 2. Match on Item Name
		if not matched_item and raw_in:
			matched_item = frappe.db.get_value("Item", {"item_name": raw_in}, ["name", "item_code", "item_name", "description", "stock_uom", "custom_msp", "brand"], as_dict=1)

		# 3. Match on Description
		if not matched_item and raw_desc:
			matched_item = frappe.db.get_value("Item", {"description": raw_desc}, ["name", "item_code", "item_name", "description", "stock_uom", "custom_msp", "brand"], as_dict=1)

		if not matched_item:
			unmatched_label = raw_ic or raw_in or raw_desc or "Row with blank item info"
			unmatched.append(unmatched_label)
			continue

		item_code = matched_item.get("name") or matched_item.get("item_code")
		item_name = matched_item.get("item_name") or raw_in or item_code
		description = matched_item.get("description") or raw_desc or item_name
		uom = matched_item.get("stock_uom") or "Nos"
		msp = flt(matched_item.get("custom_msp") or 0.0)

		# Fetch Last Sale Rate (latest submitted Sales Invoice)
		last_sale_rate = 0.0
		if header_customer:
			c_rate = frappe.db.sql("""
				SELECT sii.rate 
				FROM `tabSales Invoice Item` sii 
				JOIN `tabSales Invoice` si ON sii.parent = si.name 
				WHERE sii.item_code = %s AND si.customer = %s AND si.docstatus = 1 
				ORDER BY si.posting_date DESC, sii.creation DESC 
				LIMIT 1
			""", (item_code, header_customer))
			if c_rate and c_rate[0][0]:
				last_sale_rate = flt(c_rate[0][0])

		if not last_sale_rate:
			g_rate = frappe.db.sql("""
				SELECT rate 
				FROM `tabSales Invoice Item` 
				WHERE item_code = %s AND docstatus = 1 
				ORDER BY creation DESC 
				LIMIT 1
			""", (item_code,))
			if g_rate and g_rate[0][0]:
				last_sale_rate = flt(g_rate[0][0])

		if not last_sale_rate:
			ip_rate = frappe.db.get_value("Item Price", {"item_code": item_code, "selling": 1}, "price_list_rate")
			if ip_rate:
				last_sale_rate = flt(ip_rate)

		# Calculate Stock Balance (SUM actual_qty from tabBin)
		bin_qty = frappe.db.sql("""
			SELECT SUM(actual_qty) FROM `tabBin` WHERE item_code = %s
		""", (item_code,))
		stock_balance = flt(bin_qty[0][0]) if bin_qty and bin_qty[0][0] else 0.0

		# Calculate Reserved Stock (Pending SO Report Formax formula)
		res_stock_data = frappe.db.sql("""
			SELECT SUM(reserved_qty - IFNULL(delivered_qty, 0))
			FROM `tabStock Reservation Entry`
			WHERE docstatus = 1 AND item_code = %s
		""", (item_code,))
		reserved_stock = flt(res_stock_data[0][0]) if res_stock_data and res_stock_data[0][0] else 0.0

		# Monthly Qty
		monthly_qty = flt(row[mq_idx]) if mq_idx != -1 and row[mq_idx] else 0.0
		qty = monthly_qty if monthly_qty > 0 else 1.0

		# SOP Date
		sop_val = ""
		if sop_idx != -1 and row[sop_idx]:
			s_val = row[sop_idx]
			if isinstance(s_val, (datetime.datetime, datetime.date)):
				sop_val = s_val.strftime("%Y-%m-%d")
			elif isinstance(s_val, str):
				try:
					sop_val = str(getdate(s_val))
				except Exception:
					sop_val = s_val

		# Customer Description
		cust_desc = str(row[cd_idx]).strip() if cd_idx != -1 and row[cd_idx] else ""

		# Sales Partner
		sales_partner = ""
		if sp_idx != -1 and row[sp_idx]:
			sp_val = str(row[sp_idx]).strip()
			sp_match = frappe.db.get_value("Sales Partner", {"partner_name": sp_val}, "name") or frappe.db.get_value("Sales Partner", {"name": sp_val}, "name")
			sales_partner = sp_match or sp_val

		items.append({
			"item_code": item_code,
			"item_name": item_name,
			"description": description,
			"uom": uom,
			"qty": qty,
			"rate": last_sale_rate,
			"amount": qty * last_sale_rate,
			"stock_balance": stock_balance,
			"custom_reserved_stock": reserved_stock,
			"custom_msp": msp,
			"custom_sop_date": sop_val,
			"custom_monthly_qty": monthly_qty,
			"custom_customer_description": cust_desc,
			"custom_sales_partner": sales_partner
		})

	return {
		"customer": header_customer,
		"project": header_project,
		"items": items,
		"unmatched": unmatched
	}

def notify_rfq_submit(doc, method):
	if doc.get("custom_product_manager"):
		pm_user = doc.custom_product_manager
		if frappe.db.exists("User", pm_user):
			nlog = frappe.new_doc("Notification Log")
			nlog.subject = _("New Request for Quotation: {0}").format(doc.name)
			nlog.email_content = _("A new Request for Quotation has been submitted and is ready to be sent to a supplier.")
			nlog.for_user = pm_user
			nlog.document_type = "Request for Quotation"
			nlog.document_name = doc.name
			nlog.insert(ignore_permissions=True)
			frappe.publish_realtime('notification', nlog.subject, user=pm_user)

def after_migrate():
	# 1. Clean up legacy custom fields on RFQ header if present
	for fld in ["Request for Quotation-custom_sales_partner", "Request for Quotation-sales_partner"]:
		if frappe.db.exists("Custom Field", fld):
			try:
				frappe.delete_doc("Custom Field", fld, ignore_permissions=True)
			except Exception:
				pass
			
	# 2. Position Product Manager right after Project on RFQ
	if frappe.db.exists("Custom Field", "Request for Quotation-custom_product_manager"):
		try:
			frappe.db.set_value("Custom Field", "Request for Quotation-custom_product_manager", "insert_after", "custom_project")
		except Exception:
			pass