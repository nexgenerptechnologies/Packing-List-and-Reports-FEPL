import frappe
from frappe.model.document import Document
from frappe import _

class ShipmentTracker(Document):
	def validate(self):
		if not frappe.db.get_single_value('Packing List Settings', 'enable_shipment_tracker'):
			frappe.throw(_('Shipment Tracker is disabled in Packing List Settings.'))
		for item in self.shipment_items:
			if not item.supplier_invoice:
				frappe.throw(_("Row {0}: Supplier Invoice # is mandatory.").format(item.idx))
			if not item.bill_date:
				frappe.throw(_("Row {0}: Invoice Date is mandatory.").format(item.idx))

		# Check within the document for conflicting Invoice Dates for the same Supplier Invoice #
		invoice_dates = {}
		for item in self.shipment_items:
			if not item.supplier_invoice or not item.bill_date:
				continue
			inv = item.supplier_invoice.strip().lower()
			date = str(item.bill_date)
			if inv in invoice_dates and invoice_dates[inv] != date:
				frappe.throw(
					_("Supplier Invoice # '{0}' has conflicting Invoice Dates in this tracker ({1} vs {2}). All rows with the same invoice number must have the same date.")
					.format(item.supplier_invoice, invoice_dates[inv], date)
				)
			invoice_dates[inv] = date
			
			# Check across other documents in the database
			existing = frappe.db.sql("""
				SELECT child.parent, child.bill_date 
				FROM `tabShipment Item` child
				JOIN `tabShipment Tracker` parent ON child.parent = parent.name
				WHERE LOWER(TRIM(child.supplier_invoice)) = %s 
				  AND child.parent != %s 
				  AND child.bill_date != %s
				  AND parent.docstatus < 2
				LIMIT 1
			""", (inv, self.name or '', item.bill_date), as_dict=True)
			
			if existing:
				frappe.throw(
					_("Supplier Invoice # '{0}' already exists in Shipment Tracker '{1}' with a different Invoice Date ({2}). The date must be {3}.")
					.format(item.supplier_invoice, existing[0].parent, existing[0].bill_date, existing[0].bill_date)
				)

		# Detailed line-item strict validation on Save
		import re
		from frappe.utils import strip_html
		def normalize_text(text):
			if not text: return ""
			t = strip_html(str(text))
			t = t.replace("\xa0", " ").replace("\r", "").replace("\n", " ")
			return re.sub(r'\s+', ' ', t).strip()
			
		validation_errors = []
		
		# First pass: Link missing PO items and accumulate quantities by PO Item
		tracker_qty_totals = {}
		for item in self.shipment_items:
			if item.qty > 0:
				if not item.purchase_order_item:
					if not item.line_number:
						validation_errors.append(_("Row {0}: Item {1} is missing both Purchase Order Item and Line Number. Cannot auto-link.").format(item.idx, item.item_code))
						continue
					
					# Auto-link based on available line number fields on Purchase Order Item
					po_item_meta = frappe.get_meta("Purchase Order Item")
					where_clauses = []
					query_args = []
					
					if po_item_meta.has_field("custom_line_number"):
						where_clauses.append("poi.custom_line_number = %s")
						query_args.append(item.line_number)
					if po_item_meta.has_field("line_number"):
						where_clauses.append("poi.line_number = %s")
						query_args.append(item.line_number)
						
					if not where_clauses:
						where_clauses.append("poi.idx = %s")
						query_args.append(item.line_number)
						
					where_cond = " OR ".join(where_clauses)
					query_args.extend([item.item_code, self.supplier])
					
					po_items = frappe.db.sql(f"""
						SELECT poi.name, poi.parent 
						FROM `tabPurchase Order Item` poi
						JOIN `tabPurchase Order` po ON poi.parent = po.name
						WHERE ({where_cond}) AND poi.item_code = %s AND po.supplier = %s AND po.docstatus = 1
					""", tuple(query_args), as_dict=1)
					
					if po_items:
						item.purchase_order = po_items[0].parent
						item.purchase_order_item = po_items[0].name
					else:
						validation_errors.append(_("Row {0}: Could not find a matching Purchase Order for Line Number '{1}' and Item '{2}'.").format(item.idx, item.line_number, item.item_code))
						continue
				
				if item.purchase_order_item:
					tracker_qty_totals[item.purchase_order_item] = tracker_qty_totals.get(item.purchase_order_item, 0.0) + float(item.qty)
					
		# Ensure parent currency matches the linked Purchase Order currency
		if self.shipment_items:
			for item in self.shipment_items:
				if item.purchase_order:
					po_currency = frappe.db.get_value("Purchase Order", item.purchase_order, "currency")
					if po_currency:
						self.currency = po_currency
						break
					
		# Second pass: Perform other validations and combined quantity validation
		reported_po_qty_errors = set()
		for item in self.shipment_items:
			if item.qty > 0 and item.purchase_order_item:
				try:
					po_item = frappe.get_doc("Purchase Order Item", item.purchase_order_item)
				except frappe.DoesNotExistError:
					validation_errors.append(_("Row {0}: Purchase Order Item {1} does not exist.").format(item.idx, item.purchase_order_item))
					continue
					
				po_line_number = po_item.get("custom_line_number") or po_item.get("line_number") or str(po_item.idx)
				
				discrepancies = []

				# Supplier Invoice Date Validation
				from frappe.utils import getdate
				po_date = frappe.db.get_value("Purchase Order", item.purchase_order, "transaction_date")
				if po_date and item.bill_date:
					if getdate(item.bill_date) < getdate(po_date):
						discrepancies.append(_("Supplier Invoice Date ({0}) is before Purchase Order Date ({1})").format(item.bill_date, po_date))

				# Item Code Match
				if normalize_text(item.item_code) != normalize_text(po_item.item_code):
					discrepancies.append(_("Item Code mismatch (Expected: '{0}', Got: '{1}')").format(po_item.item_code or "", item.item_code or ""))
				
				# Item Name Match
				if normalize_text(item.item_name) != normalize_text(po_item.item_name):
					discrepancies.append(_("Item Name mismatch (Expected: '{0}', Got: '{1}')").format(po_item.item_name or "", item.item_name or ""))
				
				# Description Match
				if normalize_text(item.description) != normalize_text(po_item.description):
					discrepancies.append(_("Description mismatch (Expected: '{0}', Got: '{1}')").format(po_item.description or "", item.description or ""))
				
				# Combined Quantity Match against PO Item's pending quantity
				other_shipped = frappe.db.sql("""
					SELECT SUM(qty)
					FROM `tabShipment Item`
					WHERE purchase_order_item = %s
					  AND docstatus = 1
					  AND parent != %s
				""", (item.purchase_order_item, self.name or ""))[0][0] or 0.0
				
				pending_qty = max(0.0, min(po_item.qty - other_shipped, po_item.qty - po_item.received_qty))
				combined_qty = tracker_qty_totals.get(item.purchase_order_item, 0.0)
				
				if float(combined_qty) > float(pending_qty):
					if item.purchase_order_item not in reported_po_qty_errors:
						discrepancies.append(_("Combined quantity ({0}) for this PO Item exceeds pending amount (Pending: {1}, Row Got: {2})").format(combined_qty, pending_qty, item.qty))
						reported_po_qty_errors.add(item.purchase_order_item)
				
				# Rate Match (Precision up to 0.000001)
				if abs(float(item.rate or 0) - float(po_item.rate or 0)) > 0.000001:
					discrepancies.append(_("Rate mismatch (Expected: {0}, Got: {1})").format(po_item.rate, item.rate))
				
				# Line Number Match
				if normalize_text(item.line_number) != normalize_text(po_line_number):
					discrepancies.append(_("Line Number mismatch (Expected: '{0}', Got: '{1}')").format(po_line_number or "", item.line_number or ""))
				
				if discrepancies:
					for discrepancy in discrepancies:
						validation_errors.append(_("Row {0} (Item {1}): {2}").format(item.idx, item.item_code, discrepancy))
					
		if validation_errors:
			error_html = "<b>" + _("Validation failed for one or more items:") + "</b><br><br>"
			error_html += "<br>".join([f"\u2022 {err}" for err in validation_errors])
			frappe.throw(error_html)

	def before_submit(self):
		if not self.shipment_items or len(self.shipment_items) == 0:
			frappe.throw(_("Please fetch or add items to the shipment before submitting."))
@frappe.whitelist()
def download_template():
	import openpyxl
	from io import BytesIO
	
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Shipment Import Template"
	
	headers = [
		"Item Code", "Item Name", "Description", "Quantity", 
		"Rate", "Line Number", "Supplier Invoice #", "Invoice Date"
	]
	ws.append(headers)
	
	from openpyxl.styles import Font
	for cell in ws[1]:
		cell.font = Font(bold=True)
		
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	
	frappe.response['filename'] = "Shipment_Template.xlsx"
	frappe.response['filecontent'] = output.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def get_outstanding_po_items(supplier, purchase_orders):
	if isinstance(purchase_orders, str):
		import json
		purchase_orders = json.loads(purchase_orders)
		
	if not purchase_orders:
		return []

	# Use custom_line_number if exists, otherwise fallback to idx (sequence number)
	return frappe.db.sql("""
		SELECT 
			poi.item_code, poi.item_name, poi.description, 
			(poi.qty - poi.received_qty) as qty, poi.rate, 
			poi.custom_line_number as line_number,
			poi.parent as purchase_order, poi.name as purchase_order_item,
			po.currency, po.conversion_rate
		FROM `tabPurchase Order Item` poi
		JOIN `tabPurchase Order` po ON poi.parent = po.name
		WHERE po.supplier = %s AND po.name IN ({0}) AND po.docstatus = 1 AND poi.qty > poi.received_qty
		ORDER BY po.name, poi.idx ASC
	""".format(", ".join(["'{0}'".format(d) for d in purchase_orders])), (supplier), as_dict=1)

@frappe.whitelist()
def make_purchase_receipt(docname):
	source_doc = frappe.get_doc("Shipment Tracker", docname)
	
	if source_doc.purchase_receipt:
		frappe.throw(_("Purchase Receipt already created for this shipment: {0}").format(source_doc.purchase_receipt))
	
	target_doc = frappe.new_doc("Purchase Receipt")
	target_doc.supplier = source_doc.supplier
	
	company, currency, conversion_rate = None, None, 1.0
	if source_doc.shipment_items:
		po_data = frappe.db.get_value("Purchase Order", source_doc.shipment_items[0].purchase_order, ["company", "currency", "conversion_rate"], as_dict=1)
		if po_data:
			company, currency, conversion_rate = po_data.company, po_data.currency, po_data.conversion_rate
					
	target_doc.company = company or frappe.defaults.get_global_default("company")
	target_doc.posting_date = frappe.utils.nowdate()
	target_doc.currency = currency
	target_doc.conversion_rate = conversion_rate
	
	# Aggregate shipment tracker items by purchase_order_item
	pr_items = {}
	for item in source_doc.shipment_items:
		if item.qty > 0:
			key = item.purchase_order_item
			if key in pr_items:
				pr_items[key]["qty"] += item.qty
			else:
				pr_items[key] = {
					"item_code": item.item_code,
					"qty": item.qty,
					"rate": item.rate,
					"purchase_order": item.purchase_order,
					"purchase_order_item": item.purchase_order_item
				}
				
	for item_data in pr_items.values():
		po_item = frappe.get_doc("Purchase Order Item", item_data["purchase_order_item"])

		pr_item = target_doc.append("items", {})
		pr_item.item_code = item_data["item_code"]
		pr_item.qty = item_data["qty"]
		pr_item.rate = item_data["rate"]
		pr_item.purchase_order = item_data["purchase_order"]
		pr_item.purchase_order_item = item_data["purchase_order_item"]
		pr_item.uom = po_item.uom
		pr_item.stock_uom = po_item.stock_uom
		pr_item.conversion_factor = po_item.conversion_factor
					
	target_doc.insert()
	return target_doc.name

@frappe.whitelist()
def create_purchase_invoices(docname):
	source_doc = frappe.get_doc("Shipment Tracker", docname)
	if not source_doc.purchase_receipt:
		frappe.throw(_("Please link a Purchase Receipt first."))
		
	invoice_groups = {}
	for item in source_doc.shipment_items:
		inv_no = item.supplier_invoice
		if inv_no:
			if inv_no not in invoice_groups: invoice_groups[inv_no] = []
			invoice_groups[inv_no].append(item)
		
	created_invoices = []
	from erpnext.stock.doctype.purchase_receipt.purchase_receipt import make_purchase_invoice
	
	for inv_no, items in invoice_groups.items():
		pi = make_purchase_invoice(source_doc.purchase_receipt)
		pi.bill_no = inv_no
		pi.bill_date = items[0].bill_date or frappe.utils.nowdate()
		
		# Map purchase_order_item -> quantity for this supplier invoice group
		qty_map = {it.purchase_order_item: it.qty for it in items if it.purchase_order_item}
		
		new_items = []
		for pi_item in pi.get("items"):
			if pi_item.po_detail in qty_map:
				pi_item.qty = qty_map[pi_item.po_detail]
				if pi_item.conversion_factor:
					pi_item.stock_qty = pi_item.qty * pi_item.conversion_factor
				else:
					pi_item.stock_qty = pi_item.qty
				new_items.append(pi_item)
		
		pi.set("items", new_items)
		
		if pi.get("items"):
			pi.insert()
			created_invoices.append(pi.name)
			for row in source_doc.shipment_invoices:
				if row.bill_no == inv_no:
					row.db_set("purchase_invoice", pi.name)
					break
					
	return created_invoices

@frappe.whitelist()
def has_purchase_invoices(purchase_receipt):
	active_invoices = frappe.db.sql("""
		SELECT 1 
		FROM `tabPurchase Invoice Item` pii
		JOIN `tabPurchase Invoice` pi ON pii.parent = pi.name
		WHERE pii.purchase_receipt = %s AND pi.docstatus < 2
		LIMIT 1
	""", (purchase_receipt,))
	return len(active_invoices) > 0

@frappe.whitelist()
def fetch_from_excel(docname):
	doc = frappe.get_doc("Shipment Tracker", docname)
	if not doc.excel_file:
		frappe.throw(_("Please attach an Excel file first."))
		
	import openpyxl
	from frappe.utils import flt, getdate
	import datetime
	
	try:
		file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
		wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
		sheet = wb.active
		
		header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
		col_map = {}
		expected = {
			"item_code": ["Item Code", "item code"],
			"item_name": ["Item Name", "item name"],
			"description": ["Description", "description"],
			"qty": ["Quantity", "Qty", "quantity"],
			"rate": ["Rate", "rate"],
			"line_number": ["Line Number", "Line #", "line number"],
			"supplier_invoice": ["Purchase Invoice Number", "Supplier Invoice No.", "Invoice No", "Supplier Invoice No", "Supplier Invoice #", "supplier invoice no."],
			"bill_date": ["Date", "Invoice Date", "Purchase Invoice Date", "invoice date"]
		}
		for idx, cell in enumerate(header_row):
			if not cell: continue
			clean = str(cell).strip().lower()
			for key, aliases in expected.items():
				if any(alias.lower() == clean for alias in aliases):
					col_map[key] = idx
		
		if "item_code" not in col_map:
			frappe.throw(_("Could not find 'Item Code' column in Excel file. Please check column headers."))
			
		doc.set("shipment_items", [])
		
		grouped_items = {}
		
		for row in sheet.iter_rows(min_row=2, values_only=True):
			if not any(row): continue
			
			item_code = str(row[col_map["item_code"]]).strip() if "item_code" in col_map and row[col_map["item_code"]] else ""
			if not item_code: continue
			
			item_name = str(row[col_map["item_name"]]).strip() if "item_name" in col_map and row[col_map["item_name"]] else ""
			description = str(row[col_map["description"]]).strip() if "description" in col_map and row[col_map["description"]] else ""
			qty = flt(row[col_map["qty"]]) if "qty" in col_map else 0.0
			rate = flt(row[col_map["rate"]]) if "rate" in col_map else 0.0
			
			line_number = str(row[col_map["line_number"]]).strip() if "line_number" in col_map and row[col_map["line_number"]] else ""
			supplier_invoice = str(row[col_map["supplier_invoice"]]).strip() if "supplier_invoice" in col_map and row[col_map["supplier_invoice"]] else ""
			
			bill_date = None
			if "bill_date" in col_map:
				raw_date = row[col_map["bill_date"]]
				if isinstance(raw_date, (datetime.datetime, datetime.date)):
					# If both month and day are <= 12, Excel/openpyxl likely swapped them due to US locale parsing
					if raw_date.day <= 12 and raw_date.month <= 12:
						bill_date = f"{raw_date.year:04d}-{raw_date.day:02d}-{raw_date.month:02d}"
					else:
						bill_date = raw_date.strftime("%Y-%m-%d")
				elif isinstance(raw_date, str):
					val = raw_date.strip()
					parsed = False
					for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d"):
						try:
							parsed_dt = datetime.datetime.strptime(val, fmt)
							bill_date = parsed_dt.strftime("%Y-%m-%d")
							parsed = True
							break
						except ValueError:
							continue
					if not parsed:
						try:
							bill_date = getdate(val).strftime("%Y-%m-%d")
						except:
							pass
			
			# Grouping Key (Club only if item, line number, and supplier invoice are exactly the same)
			key = (item_code.lower(), line_number.lower(), supplier_invoice.lower())
			
			if key in grouped_items:
				grouped_items[key]["qty"] += qty
				if not grouped_items[key]["bill_date"] and bill_date:
					grouped_items[key]["bill_date"] = bill_date
			else:
				grouped_items[key] = {
					"item_code": item_code,
					"item_name": item_name,
					"description": description,
					"qty": qty,
					"rate": rate,
					"line_number": line_number,
					"supplier_invoice": supplier_invoice,
					"bill_date": bill_date
				}
				
		# Append the aggregated items to the document
		for item_data in grouped_items.values():
			child = doc.append("shipment_items", {})
			child.item_code = item_data["item_code"]
			child.item_name = item_data["item_name"]
			child.description = item_data["description"]
			child.qty = item_data["qty"]
			child.rate = item_data["rate"]
			child.line_number = item_data["line_number"]
			child.supplier_invoice = item_data["supplier_invoice"]
			child.bill_date = item_data["bill_date"]
						
		doc.save()
		return "Success"
		
	except frappe.ValidationError:
		raise
	except Exception as e:
		frappe.throw(f"Failed to parse Excel: {str(e)}")

@frappe.whitelist()
def add_shipment_remark(docname, remark):
	if not remark or not remark.strip():
		return
		
	doc = frappe.get_doc("Shipment Tracker", docname)
	
	from frappe.utils import format_datetime, now_datetime
	user_fullname = frappe.db.get_value("User", frappe.session.user, "full_name") or frappe.session.user
	timestamp = format_datetime(now_datetime(), "dd-MM-yyyy hh:mm a")
	
	new_remark_entry = f"[{timestamp}] {user_fullname}:\n{remark.strip()}"
	
	if doc.remarks:
		doc.remarks = f"{new_remark_entry}\n\n----------------------------------------\n\n{doc.remarks}"
	else:
		doc.remarks = new_remark_entry
		
	doc.db_set("remarks", doc.remarks)
	return "Success"
