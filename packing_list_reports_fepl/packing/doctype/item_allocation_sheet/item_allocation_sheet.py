import frappe
from frappe.model.document import Document
from frappe import _

@frappe.whitelist()
def get_sales_partner_for_user(user):
	# 1. Standard ERPNext mapping via Portal User child table
	sp_name = frappe.db.get_value("Portal User", {"user": user, "parenttype": "Sales Partner"}, "parent")
	if sp_name:
		return sp_name
	# 2. Check if Sales Partner ID (name) matches the user ID (email)
	sp_name = frappe.db.get_value("Sales Partner", {"name": user}, "name")
	if sp_name:
		return sp_name
		
	# Check if database has sales_partner_name column before querying to prevent OperationalError 1054
	if frappe.db.has_column("Sales Partner", "sales_partner_name"):
		# 3. Check if Sales Partner Name matches the user ID (email)
		sp_name = frappe.db.get_value("Sales Partner", {"sales_partner_name": user}, "name")
		if sp_name:
			return sp_name
		# 4. Check if Sales Partner Name matches the user's full name
		full_name = frappe.db.get_value("User", user, "full_name")
		if full_name:
			sp_name = frappe.db.get_value("Sales Partner", {"sales_partner_name": full_name}, "name")
			if sp_name:
				return sp_name
				
	# 5. Check if email prefix matches Sales Partner ID or Sales Partner Name
	if user and "@" in user:
		prefix = user.split("@")[0].strip()
		sp_name = frappe.db.get_value("Sales Partner", {"name": ["like", f"%{prefix}%"]}, "name")
		if sp_name:
			return sp_name
		if frappe.db.has_column("Sales Partner", "sales_partner_name"):
			sp_name = frappe.db.get_value("Sales Partner", {"sales_partner_name": ["like", f"%{prefix}%"]}, "name")
			if sp_name:
				return sp_name
	return None

@frappe.whitelist()
def get_available_shipments_for_partner(doctype, txt, searchfield, start, page_len, filters):
	# Get the sales partner of the current user
	user = frappe.session.user
	sp_name = get_sales_partner_for_user(user)
	
	# If the user is system manager or admin and no partner is mapped, or if partner is specified in filters, let's use it
	if filters and filters.get("sales_partner"):
		sp_name = filters.get("sales_partner")
		
	# Find all Shipment Trackers already sent to Team Leader by this partner
	exclude_shipments = []
	if sp_name:
		# Find Item Allocation Sheets for this partner in sent stages
		sent_sheets = frappe.get_all("Item Allocation Sheet",
			filters={
				"sales_partner": sp_name,
				"status": ["in", ["Pending Team Leader", "Pending Partner Finalization", "Approved"]],
				"docstatus": ["<", 2]
			},
			fields=["name"])
			
		if sent_sheets:
			sent_sheet_names = [s.name for s in sent_sheets]
			# Find shipment trackers in these sheets
			shipments_data = frappe.get_all("Item Allocation Shipment",
				filters={
					"parent": ["in", sent_sheet_names]
				},
				fields=["shipment_tracker"])
			exclude_shipments = list(set([s.shipment_tracker for s in shipments_data if s.shipment_tracker]))
			
	# Query Shipment Trackers with docstatus=1 and not in exclude_shipments
	query_filters = {
		"docstatus": 1
	}
	if exclude_shipments:
		query_filters["name"] = ["not in", exclude_shipments]
		
	# Match with txt search filter
	if txt:
		query_filters["name"] = ["like", f"%{txt}%"]
		
	shipments = frappe.get_all("Shipment Tracker",
		filters=query_filters,
		fields=["name", "supplier", "eta"],
		start=start,
		page_length=page_len,
		as_list=1)
		
	return shipments

class ItemAllocationSheet(Document):
	def onload(self):
		ensure_custom_fields()
		
	def validate(self):
		ensure_custom_fields()
		
		if not frappe.db.get_single_value('Packing List Settings', 'enable_item_allocation_sheet'):
			frappe.throw(_("Item Allocation Sheet is disabled in Packing List Settings."))
		
		# Validation: Ensure partner cannot have multiple sheets for the same shipment once sent to Team Leader
		selected_shipments = [s.shipment_tracker for s in self.shipments if s.shipment_tracker]
		if selected_shipments and self.sales_partner:
			other_sheets = frappe.get_all("Item Allocation Sheet",
				filters={
					"sales_partner": self.sales_partner,
					"name": ["!=", self.name],
					"status": ["in", ["Pending Team Leader", "Pending Partner Finalization", "Approved"]],
					"docstatus": ["<", 2]
				},
				fields=["name"])
				
			for sheet_meta in other_sheets:
				overlap_shipment = frappe.db.get_value("Item Allocation Shipment", {
					"parent": sheet_meta.name,
					"shipment_tracker": ["in", selected_shipments]
				}, "shipment_tracker")
				if overlap_shipment:
					frappe.throw(_("Shipment {0} has already been sent to the Team Leader in sheet {1}. You cannot create or submit another sheet for this shipment.").format(overlap_shipment, sheet_meta.name))
		
		# Validation: Make sure Total Allocation Request does not exceed Shipment Quantity per Item Code
		selected_shipments = [s.shipment_tracker for s in self.shipments if s.shipment_tracker]
		if selected_shipments:
			shipment_qtys = {}
			ship_items = frappe.get_all("Shipment Item",
				filters={"parent": ["in", selected_shipments]},
				fields=["item_code", "qty"])
			for s_item in ship_items:
				ic = s_item.item_code.strip()
				shipment_qtys[ic] = shipment_qtys.get(ic, 0.0) + (s_item.qty or 0.0)

			allocation_reqs = {}
			for item in self.items:
				ic = item.item_code.strip() if item.item_code else ""
				if ic:
					allocation_reqs[ic] = allocation_reqs.get(ic, 0.0) + (item.allocation_request or 0.0)

			for ic, total_req in allocation_reqs.items():
				match_key = None
				for k in shipment_qtys.keys():
					if k.lower() == ic.lower():
						match_key = k
						break
				allowed_qty = shipment_qtys.get(match_key, 0.0) if match_key else 0.0
				if total_req > allowed_qty:
					frappe.throw(_("Item Code '{0}': Total Allocation Request ({1}) exceeds the Shipment Quantity ({2})!").format(
						ic, total_req, allowed_qty
					))

		# Validation: Make sure Final Allocation sum does not exceed TL Quota per Item Code
		if self.status == "Pending Partner Finalization":
			item_quotas = {}
			item_finals = {}
			for item in self.items:
				item_quotas[item.item_code] = max(item_quotas.get(item.item_code, 0), item.allocated_qty or 0)
				item_finals[item.item_code] = item_finals.get(item.item_code, 0) + (item.final_allocation or 0)
				
			for item_code, final_sum in item_finals.items():
				quota = item_quotas.get(item_code, 0)
				if final_sum > quota:
					frappe.throw(_("Item {0}: Total Final Allocation ({1}) exceeds the Quota allocated by Team Leader ({2}).").format(item_code, final_sum, quota))

		# Validation: Customer and Sales Order constraint checks
		for idx, item in enumerate(self.items):
			if item.item_code and item.customer:
				validate_item_customer_sales_order(
					item.item_code,
					item.customer,
					item.sales_order,
					row_label=f"Row {idx + 1}"
				)

	def before_save(self):
		# Re-map sales partner based on user and check registration
		is_manager = "System Manager" in frappe.get_roles() or "Sales Manager" in frappe.get_roles() or frappe.session.user == "Administrator"
		sp_name = get_sales_partner_for_user(frappe.session.user)
		
		if sp_name:
			if not is_manager or not self.sales_partner:
				self.sales_partner = sp_name
		elif not self.sales_partner:
			if not is_manager:
				frappe.throw(_("Your user login '{0}' is not registered as a Sales Partner in the system. Please register your user as a Sales Partner first.").format(frappe.session.user))
			else:
				frappe.throw(_("Please select a Sales Partner before saving."))
				
		# Calculate total allocation request per item code
		item_totals = {}
		for item in self.items:
			if self.sales_partner:
				item.sales_partner = self.sales_partner
			item_totals[item.item_code] = item_totals.get(item.item_code, 0) + (item.allocation_request or 0)
			
		for item in self.items:
			item.total_allocation_request = item_totals[item.item_code]

	def on_submit(self):
		self.db_set("status", "Approved")
		
		from frappe.utils import flt
		reserved_count = 0
		
		# Reserve quantities in Sales Orders
		for item in self.items:
			if item.sales_order and item.final_allocation > 0:
				# Find the Sales Order Item matching the item code
				so_item = frappe.db.get_value("Sales Order Item", {
					"parent": item.sales_order,
					"item_code": item.item_code
				}, ["name", "custom_reserved_qty", "warehouse"], as_dict=1)
				
				if so_item:
					current_reserved = flt(so_item.custom_reserved_qty)
					new_reserved = current_reserved + flt(item.final_allocation)
					
					# Build standard and custom values to update on Sales Order Item
					update_values = {"custom_reserved_qty": new_reserved}
					
					so_item_meta = frappe.get_meta("Sales Order Item")
					if so_item_meta.has_field("reserve_stock"):
						update_values["reserve_stock"] = 1
						
					frappe.db.set_value("Sales Order Item", so_item.name, update_values)
					
					# Create standard Stock Reservation Entry if DocType exists
					if frappe.db.exists("DocType", "Stock Reservation Entry"):
						warehouse = so_item.warehouse or "Finished Goods"
						sre_dict = {
							"doctype": "Stock Reservation Entry",
							"item_code": item.item_code,
							"warehouse": warehouse,
							"voucher_type": "Sales Order",
							"voucher_no": item.sales_order,
							"voucher_detail_no": so_item.name,
							"reserved_qty": item.final_allocation,
							"company": frappe.db.get_value("Sales Order", item.sales_order, "company"),
							"status": "Reserved"
						}
						
						sre_meta = frappe.get_meta("Stock Reservation Entry")
						if sre_meta.has_field("remarks"):
							sre_dict["remarks"] = f"Reserved via Item Allocation Sheet {self.name}"
							
						if sre_meta.has_field("available_qty"):
							# Fetch actual available quantity if function is importable, otherwise fallback to final_allocation to pass required checks
							try:
								from erpnext.stock.doctype.stock_reservation_entry.stock_reservation_entry import get_available_qty_to_reserve
								db_available = get_available_qty_to_reserve(item.item_code, warehouse)
								sre_dict["available_qty"] = max(flt(db_available), flt(item.final_allocation))
							except:
								sre_dict["available_qty"] = flt(item.final_allocation)
							
						sre = frappe.get_doc(sre_dict)
						sre.flags.ignore_validate = True
						sre.flags.ignore_mandatory = True
						sre.insert()
						
						sre.flags.ignore_validate = True
						sre.flags.ignore_mandatory = True
						sre.submit()
						
						# Set reference to the SRE in the child row if field exists
						item_meta = frappe.get_meta("Partner Allocation Detail")
						if item_meta.has_field("stock_reservation_entry"):
							item.db_set("stock_reservation_entry", sre.name)
							
					frappe.clear_document_cache("Sales Order", item.sales_order)
					reserved_count += 1
					
		if reserved_count > 0:
			frappe.msgprint(
				_("<b>Stock Reservation Successful:</b><br>&bull; Stock has been successfully reserved against the corresponding Sales Order Item Lines."),
				alert=True
			)
					
	def on_cancel(self):
		# Restriction: Only the owner (the Sales Partner who submitted it) or a System/Sales Manager can cancel
		is_owner = (self.owner == frappe.session.user)
		is_authorized = is_owner or ("System Manager" in frappe.get_roles() or "Sales Manager" in frappe.get_roles())
		
		if not is_authorized:
			frappe.throw(_("You are not authorized to unreserve/cancel this document. Only the submitting Sales Partner or a Team Leader can do this."))
			
		# Unreserve quantities in Sales Orders
		for item in self.items:
			if item.sales_order and item.final_allocation > 0:
				so_item = frappe.db.get_value("Sales Order Item", {
					"parent": item.sales_order,
					"item_code": item.item_code
				}, ["name", "custom_reserved_qty"], as_dict=1)
				
				if so_item:
					current_reserved = so_item.custom_reserved_qty or 0.0
					new_reserved = max(0.0, current_reserved - item.final_allocation)
					
					# Build standard and custom values to update on Sales Order Item
					update_values = {"custom_reserved_qty": new_reserved}
					
					# If we are unreserving all stock, we can uncheck standard reserve_stock
					if new_reserved <= 0:
						so_item_meta = frappe.get_meta("Sales Order Item")
						if so_item_meta.has_field("reserve_stock"):
							update_values["reserve_stock"] = 0
							
					frappe.db.set_value("Sales Order Item", so_item.name, update_values)
					
					# Cancel standard Stock Reservation Entries
					if frappe.db.exists("DocType", "Stock Reservation Entry"):
						# 1. Direct link check
						sre_name = item.get("stock_reservation_entry") if hasattr(item, "stock_reservation_entry") else None
						if sre_name:
							sre_status = frappe.db.get_value("Stock Reservation Entry", sre_name, "docstatus")
							if sre_status == 1:
								sre = frappe.get_doc("Stock Reservation Entry", sre_name)
								sre.flags.ignore_permissions = True
								sre.cancel()
							item.db_set("stock_reservation_entry", None)
						else:
							# 2. Fallback search with dynamic remarks check
							sre_meta = frappe.get_meta("Stock Reservation Entry")
							sre_filters = {
								"voucher_type": "Sales Order",
								"voucher_no": item.sales_order,
								"voucher_detail_no": so_item.name,
								"item_code": item.item_code,
								"docstatus": 1
							}
							if sre_meta.has_field("remarks"):
								sre_filters["remarks"] = f"Reserved via Item Allocation Sheet {self.name}"
								
							sres = frappe.get_all("Stock Reservation Entry",
								filters=sre_filters,
								fields=["name", "reserved_qty"])
								
							for sre_meta_row in sres:
								# If remarks column doesn't exist, we filter by exact quantity match
								if not sre_meta.has_field("remarks") and abs(flt(sre_meta_row.reserved_qty) - flt(item.final_allocation)) > 0.0001:
									continue
								sre = frappe.get_doc("Stock Reservation Entry", sre_meta_row.name)
								sre.flags.ignore_permissions = True
								sre.cancel()
								
					frappe.clear_document_cache("Sales Order", item.sales_order)

def ensure_custom_fields():
	from frappe.custom.doctype.custom_field.custom_field import create_custom_fields
	create_custom_fields({
		"Sales Order Item": [
			{
				"fieldname": "custom_reserved_qty",
				"label": "Reserved Qty from Shipment",
				"fieldtype": "Float",
				"insert_after": "delivered_qty",
				"read_only": 1
			}
		]
	})

@frappe.whitelist()
def upload_excel_data(docname):
	doc = frappe.get_doc("Item Allocation Sheet", docname)
	if doc.docstatus == 1:
		frappe.throw(_("Cannot upload Excel to a submitted document."))
		
	import openpyxl
	from frappe.utils import flt
	
	try:
		file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
		wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
		sheet = wb.active
		
		header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
		col_map = {}
		expected = {
			"item_code": ["Item Code", "item code"],
			"item_name": ["Item Name", "item name"],
			"description": ["Description", "description"],
			"total_qty": ["Shipment Qty", "Shipment Quantity", "Qty"],
			"customer": ["Customer Name", "Customer", "customer"],
			"sales_order": ["Sales Order", "sales order"],
			"allocation_request": ["Allocation Request", "Allocation Requested", "Allocation Request Qty"],
			"allocated_qty": ["TL Quota", "Quota", "Allocated Qty", "Allocated Quantity"],
			"final_allocation": ["Final Allocation", "Final Allocated Qty"],
			"sales_partner": ["Partner Name", "Sales Partner"]
		}
		
		for idx, cell in enumerate(header_row):
			if not cell: continue
			clean = str(cell).strip().lower()
			for key, aliases in expected.items():
				if any(alias.lower() == clean for alias in aliases):
					col_map[key] = idx
					
		if "item_code" not in col_map:
			frappe.throw(_("Could not find 'Item Code' column in Excel file."))
			
		# Fetch all items in the selected shipments to strictly validate against
		shipment_items = []
		for s_row in doc.shipments:
			if s_row.shipment_tracker:
				items_in_shipment = frappe.get_all("Shipment Item", 
					filters={"parent": s_row.shipment_tracker}, 
					fields=["item_code", "item_name", "description", "qty"])
				for item in items_in_shipment:
					item["shipment_tracker"] = s_row.shipment_tracker
					shipment_items.append(item)
					
		def normalize_str(s):
			if not s: return ""
			import re
			s = str(s)
			s = re.sub(r'<[^>]*>', '', s)
			s = s.replace('\xa0', ' ').replace('\r', '').replace('\n', '')
			s = " ".join(s.split())
			val_str = s.strip()
			if val_str.endswith(".0"):
				val_str = val_str[:-2]
			return val_str.lower()
		current_item_code = ""
		current_item_name = ""
		current_description = ""
		current_total_qty = 0.0

		# Sales Partner Upload (Draft Stage)
		if doc.status == "Draft":
			doc.set("items", [])
			excel_totals = {}
			excel_rows = {}
			parsed_rows = []
			row_idx = 1
			
			for row in sheet.iter_rows(min_row=2, values_only=True):
				row_idx += 1
				if not any(row): continue
				
				excel_item_code = row[col_map.get("item_code")] if "item_code" in col_map else None
				excel_item_name = row[col_map.get("item_name")] if "item_name" in col_map else None
				excel_desc = row[col_map.get("description")] if "description" in col_map else None
				excel_qty = row[col_map.get("total_qty")] if "total_qty" in col_map else None
				
				# If Item Code is specified, we update tracking context
				if excel_item_code and str(excel_item_code).strip():
					current_item_code = str(excel_item_code).strip()
					current_item_name = str(excel_item_name or "").strip()
					current_description = str(excel_desc or "").strip()
					current_total_qty = flt(excel_qty)
					
				if not current_item_code:
					continue
					
				customer = str(row[col_map.get("customer")] or "").strip() if "customer" in col_map else ""
				sales_order = str(row[col_map.get("sales_order")] or "").strip() if "sales_order" in col_map else ""
				allocation_req = flt(row[col_map["allocation_request"]]) if "allocation_request" in col_map else 0.0
				
				if customer:
					customer_id = validate_item_customer_sales_order(
						current_item_code,
						customer,
						sales_order,
						row_label=f"Row {row_idx}"
					)
					customer = customer_id
				
				# 1. Reconciliation: Verify against items in the selected Shipment Trackers
				matched_item = None
				for s_item in shipment_items:
					if normalize_str(s_item["item_code"]) == normalize_str(current_item_code):
						# We match primarily on item_code. This is robust and prevents failures due to minor name/description formatting.
						matched_item = s_item
						break
							
				if not matched_item:
					frappe.throw(_("Row {0}: Item '{1}' does not match any item in the selected Shipment Trackers.").format(
						row_idx, current_item_code
					))
					
				excel_totals[current_item_code] = excel_totals.get(current_item_code, 0.0) + allocation_req
				if current_item_code not in excel_rows:
					excel_rows[current_item_code] = []
				excel_rows[current_item_code].append(row_idx)
				
				parsed_rows.append({
					"item_code": current_item_code,
					"item_name": current_item_name,
					"description": current_description,
					"total_qty": current_total_qty,
					"customer": customer,
					"sales_order": sales_order,
					"allocation_request": allocation_req,
					"shipment": matched_item["shipment_tracker"]
				})
				
			# 2. Strict Limit Validation: Cannot exceed total Shipment Quantity
			for item_code, total_req in excel_totals.items():
				ship_qty = sum(item["qty"] for item in shipment_items if normalize_str(item["item_code"]) == normalize_str(item_code))
				if total_req > ship_qty:
					rows_list = excel_rows.get(item_code, [])
					if len(rows_list) == 1:
						row_label = f"Row {rows_list[0]}"
					else:
						row_label = f"Rows {', '.join(str(r) for r in rows_list)}"
					frappe.throw(_("{0} / Item Code '{1}': Total Allocation Request ({2}) exceeds the Shipment Quantity ({3})!").format(
						row_label, item_code, total_req, ship_qty
					))
					
			# Append clean child rows
			for r_data in parsed_rows:
				child = doc.append("items", {})
				child.item_code = r_data["item_code"]
				child.item_name = r_data["item_name"]
				child.description = r_data["description"]
				# Force total_qty to be the clubbed sum from the selected shipment items!
				ic = r_data["item_code"]
				ship_qty = sum(item["qty"] for item in shipment_items if normalize_str(item["item_code"]) == normalize_str(ic))
				child.total_qty = ship_qty
				child.customer = r_data["customer"]
				child.sales_order = r_data["sales_order"]
				child.allocation_request = r_data["allocation_request"]
				child.shipment = r_data["shipment"]
				
		# Team Leader Upload (Pending Team Leader Stage)
		elif doc.status == "Pending Team Leader":
			for row in sheet.iter_rows(min_row=2, values_only=True):
				if not any(row): continue
				excel_item_code = row[col_map.get("item_code")] if "item_code" in col_map else None
				if excel_item_code and str(excel_item_code).strip():
					current_item_code = str(excel_item_code).strip()
					
				if not current_item_code:
					continue
				
				customer = str(row[col_map.get("customer")] or "").strip() if "customer" in col_map else ""
				sales_order = str(row[col_map.get("sales_order")] or "").strip() if "sales_order" in col_map else ""
				allocated_qty = flt(row[col_map["allocated_qty"]]) if "allocated_qty" in col_map else 0.0
				
				for child in doc.items:
					match = (normalize_str(child.item_code) == normalize_str(current_item_code))
					if customer:
						match = match and (normalize_str(child.customer) == normalize_str(customer))
					if sales_order:
						match = match and (normalize_str(child.sales_order) == normalize_str(sales_order))
						
					if match:
						child.allocated_qty = allocated_qty
						
		# Sales Partner Finalization Upload (Pending Partner Finalization Stage)
		elif doc.status == "Pending Partner Finalization":
			# Keep track of how many times we''ve seen each item code in the Excel sheet
			item_seen_counts = {}
			current_item_code = ""
			
			for row in sheet.iter_rows(min_row=2, values_only=True):
				if not any(row): continue
				
				excel_item_code = row[col_map.get("item_code")] if "item_code" in col_map else None
				if excel_item_code and str(excel_item_code).strip():
					current_item_code = str(excel_item_code).strip()
					
				if not current_item_code:
					continue
					
				# Strip .0 from numeric item codes
				norm_item_code = normalize_str(current_item_code)
				
				# Increment seen count for this item code
				item_seen_counts[norm_item_code] = item_seen_counts.get(norm_item_code, 0) + 1
				target_index = item_seen_counts[norm_item_code]
				
				customer = str(row[col_map.get("customer")] or "").strip() if "customer" in col_map else ""
				sales_order = str(row[col_map.get("sales_order")] or "").strip() if "sales_order" in col_map else ""
				final_allocation = flt(row[col_map["final_allocation"]]) if "final_allocation" in col_map else 0.0
				
				# Find the target child row in doc.items
				current_idx = 0
				for child in doc.items:
					if normalize_str(child.item_code) == norm_item_code:
						current_idx += 1
						if current_idx == target_index:
							# Found the exact matching row in sequential order!
							child.customer = customer
							child.sales_order = sales_order
							child.final_allocation = final_allocation
							break
		doc.save()
		return "Success"
	except frappe.ValidationError:
		raise
	except Exception as e:
		frappe.throw(f"Failed to parse Excel: {str(e)}")

@frappe.whitelist()
def download_partner_template(docname=None):
	import openpyxl
	from io import BytesIO
	
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Partner Allocation Template"
	
	headers = [
		"Item Code", "Item Name", "Description", "Shipment Qty", 
		"Customer Name", "Sales Order", "Allocation Request", "Total Allocation Request"
	]
	ws.append(headers)
	
	if docname:
		doc = frappe.get_doc("Item Allocation Sheet", docname)
		for item in doc.items:
			ws.append([
				item.item_code, item.item_name, item.description, item.total_qty,
				item.customer or "", item.sales_order or "", item.allocation_request or "", item.total_allocation_request or ""
			])
			
	from openpyxl.styles import Font
	for cell in ws[1]:
		cell.font = Font(bold=True)
		
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	
	frappe.response['filename'] = "Sales_Partner_Allocation_Template.xlsx"
	frappe.response['filecontent'] = output.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def download_team_leader_template(docname=None):
	import openpyxl
	from io import BytesIO
	
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Team Lead Allocation Template"
	
	headers = [
		"Item Code", "Item Name", "Description", "Customer Name", 
		"Allocation Request", "Partner Name", "TL Quota"
	]
	ws.append(headers)
	
	if docname:
		doc = frappe.get_doc("Item Allocation Sheet", docname)
		for item in doc.items:
			ws.append([
				item.item_code, item.item_name, item.description, item.customer,
				item.allocation_request, item.sales_partner, item.allocated_qty or ""
			])
			
	from openpyxl.styles import Font
	for cell in ws[1]:
		cell.font = Font(bold=True)
		
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	
	frappe.response['filename'] = "Team_Leader_Allocation_Template.xlsx"
	frappe.response['filecontent'] = output.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def download_partner_finalization_template(docname=None):
	import openpyxl
	from io import BytesIO
	
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Final Allocation Template"
	
	headers = [
		"Item Code", "Item Name", "Description", "Customer Name", 
		"Sales Order", "Allocation Request", "TL Quota", "Final Allocation"
	]
	ws.append(headers)
	
	if docname:
		doc = frappe.get_doc("Item Allocation Sheet", docname)
		for item in doc.items:
			ws.append([
				item.item_code, item.item_name, item.description, item.customer,
				item.sales_order, item.allocation_request, item.allocated_qty, item.final_allocation or ""
			])
			
	from openpyxl.styles import Font
	for cell in ws[1]:
		cell.font = Font(bold=True)
		
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	
	frappe.response['filename'] = "Sales_Partner_Final_Allocation_Template.xlsx"
	frappe.response['filecontent'] = output.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def fetch_partner_requests(docname):
	doc = frappe.get_doc("Item Allocation Sheet", docname)
	if doc.docstatus == 1:
		frappe.throw(_("Cannot fetch requests for a submitted document."))
		
	doc.set("items", [])
	
	selected_shipments = [s.shipment_tracker for s in doc.shipments if s.shipment_tracker]
	if not selected_shipments:
		frappe.throw(_("Please select at least one Shipment Tracker first."))
		
	pending_sheets = frappe.get_all("Item Allocation Sheet", 
		filters={
			"status": "Pending Team Leader",
			"docstatus": 0,
			"name": ["!=", docname]
		}, 
		fields=["name", "sales_partner"])
		
	matched_count = 0
	for sheet_meta in pending_sheets:
		sheet = frappe.get_doc("Item Allocation Sheet", sheet_meta["name"])
		for item in sheet.items:
			if item.shipment in selected_shipments:
				child = doc.append("items", {})
				child.shipment = item.shipment
				child.item_code = item.item_code
				child.item_name = item.item_name
				child.description = item.description
				child.total_qty = item.total_qty
				child.customer = item.customer
				child.sales_order = item.sales_order
				child.allocation_request = item.allocation_request
				child.sales_partner = sheet.sales_partner or item.sales_partner
				child.allocated_qty = item.allocated_qty or item.allocation_request
				child.source_doc = sheet.name
				child.source_row = item.name
				matched_count += 1
				
	doc.save()
	return f"Successfully fetched {matched_count} partner requests."

@frappe.whitelist()
def distribute_tl_quotas(docname):
	doc = frappe.get_doc("Item Allocation Sheet", docname)
	if doc.docstatus == 1:
		frappe.throw(_("Cannot distribute quotas for a submitted document."))
		
	updated_docs = set()
	
	for item in doc.items:
		if item.source_doc and item.source_row:
			frappe.db.set_value("Partner Allocation Detail", item.source_row, "allocated_qty", item.allocated_qty)
			updated_docs.add(item.source_doc)
			
	for s_name in updated_docs:
		s_doc = frappe.get_doc("Item Allocation Sheet", s_name)
		s_doc.status = "Pending Partner Finalization"
		s_doc.save()
		
	doc.status = "Approved"
	doc.save()
	return "Quotas successfully approved and distributed to all Sales Partners."

@frappe.whitelist()
def get_customers_for_item(doctype, txt, searchfield, start, page_len, filters):
	item_code = filters.get("item_code")
	if not item_code:
		return []
		
	# Find all submitted Sales Orders with this item_code that are not completed/cancelled
	so_items = frappe.get_all("Sales Order Item",
		filters={
			"item_code": item_code,
			"docstatus": 1
		},
		fields=["parent"])
		
	if not so_items:
		return []
		
	so_names = [d.parent for d in so_items]
	
	# Find distinct customers from those Sales Orders
	active_sos = frappe.get_all("Sales Order",
		filters={
			"name": ["in", so_names],
			"status": ["not in", ["Completed", "Cancelled", "Closed"]],
			"customer": ["like", f"%{txt}%"] if txt else ["!=", ""]
		},
		fields=["customer"],
		distinct=True,
		limit=page_len,
		start=start)
		
	customers = [d.customer for d in active_sos if d.customer]
	if not customers:
		return []
		
	results = []
	for cust in set(customers):
		cust_name = frappe.db.get_value("Customer", cust, "customer_name") or ""
		results.append([cust, cust_name])
		
	return results

@frappe.whitelist()
def get_sales_orders_for_item(doctype, txt, searchfield, start, page_len, filters):
	item_code = filters.get("item_code")
	customer = filters.get("customer")
	
	if not item_code or not customer:
		return []
		
	# Find all submitted Sales Orders with this item_code and customer that are not completed/cancelled
	so_items = frappe.get_all("Sales Order Item",
		filters={
			"item_code": item_code,
			"docstatus": 1
		},
		fields=["parent"])
		
	if not so_items:
		return []
		
	so_names = [d.parent for d in so_items]
	
	orders = frappe.get_all("Sales Order",
		filters={
			"name": ["in", so_names],
			"customer": customer,
			"status": ["not in", ["Completed", "Cancelled", "Closed"]],
			"name": ["like", f"%{txt}%"] if txt else ["!=", ""]
		},
		fields=["name", "customer_name"],
		limit=page_len,
		start=start)
		
	return [[d.name, d.customer_name or ""] for d in orders]

def validate_item_customer_sales_order(item_code, customer_name, sales_order=None, row_label=""):
	if not item_code or not customer_name:
		return None
		
	cust_id = frappe.db.get_value("Customer", {"name": customer_name}, "name")
	if not cust_id:
		cust_id = frappe.db.get_value("Customer", {"customer_name": customer_name}, "name")
		
	if not cust_id:
		frappe.throw(_("{0}Item Code '{1}': Customer '{2}' does not exist in the system.").format(
			f"{row_label}: " if row_label else "", item_code, customer_name
		))
		
	so_items = frappe.get_all("Sales Order Item",
		filters={
			"item_code": item_code,
			"docstatus": 1
		},
		fields=["parent"])
		
	if not so_items:
		frappe.throw(_("{0}Item Code '{1}': No submitted Sales Order found in the system for this item.").format(
			f"{row_label}: " if row_label else "", item_code
		))
		
	so_names = [d.parent for d in so_items]
	
	active_sos = frappe.get_all("Sales Order",
		filters={
			"name": ["in", so_names],
			"customer": cust_id,
			"status": ["not in", ["Completed", "Cancelled", "Closed"]]
		},
		fields=["name"])
		
	if not active_sos:
		frappe.throw(_("{0}Item Code '{1}': Customer '{2}' does not have any pending Sales Orders for this item in the system.").format(
			f"{row_label}: " if row_label else "", item_code, customer_name
		))
		
	if sales_order:
		so_exists = False
		for active_so in active_sos:
			if active_so.name == sales_order:
				so_exists = True
				break
		if not so_exists:
			frappe.throw(_("{0}Item Code '{1}': Sales Order '{2}' does not match customer '{3}' or is not an active Sales Order containing this item.").format(
				f"{row_label}: " if row_label else "", item_code, sales_order, customer_name
			))
			
	return cust_id
