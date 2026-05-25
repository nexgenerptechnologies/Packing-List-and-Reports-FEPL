# -*- coding: utf-8 -*-
# Copyright (c) 2026, Administrator and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from frappe import _
from frappe.utils import flt
import openpyxl
from io import BytesIO

def normalize_str(val):
	if val is None:
		return ""
	val_str = str(val).strip()
	if val_str.endswith(".0"):
		val_str = val_str[:-2]
	return val_str.lower().replace(" ", "").replace("-", "").replace("_", "")

def customers_match(c1, c2):
	if not c1 and not c2:
		return True
	if not c1 or not c2:
		return False
	
	n1 = normalize_str(c1)
	n2 = normalize_str(c2)
	if n1 == n2:
		return True
		
	# Resolve both to customer_name for standard comparison if one is primary key ID and other is display name
	has_field = False
	try:
		has_field = frappe.get_meta("Customer").get_field("customer_name") is not None
	except Exception:
		pass
		
	if has_field:
		name1 = frappe.db.get_value("Customer", c1, "customer_name") or c1
		name2 = frappe.db.get_value("Customer", c2, "customer_name") or c2
		return normalize_str(name1) == normalize_str(name2)
	return False

def partners_match(p1, p2):
	if not p1 and not p2:
		return True
	if not p1 or not p2:
		return False
	
	n1 = normalize_str(p1)
	n2 = normalize_str(p2)
	if n1 == n2:
		return True
		
	# Resolve both to sales_partner_name for standard comparison if one is primary key ID and other is display name
	has_field = False
	try:
		has_field = frappe.get_meta("Sales Partner").get_field("sales_partner_name") is not None
	except Exception:
		pass
		
	if has_field:
		name1 = frappe.db.get_value("Sales Partner", p1, "sales_partner_name") or p1
		name2 = frappe.db.get_value("Sales Partner", p2, "sales_partner_name") or p2
		return normalize_str(name1) == normalize_str(name2)
	return False

class TeamLeaderAllocationSheet(Document):
	def validate(self):
		if not frappe.db.get_single_value('Packing List Settings', 'enable_team_leader_allocation_sheet'):
			frappe.throw(_('Team Leader Allocation Sheet is disabled in Packing List Settings.'))
		
		# Quota validation runs ONLY when the document status is set to "Approved".
		# During draft stages (e.g. Fetching requests), Sales Partners are allowed to have requested quantities
		# whose sum exceeds shipment qty, and the Team Leader can fetch them without any validation crash.
		if self.status == "Approved":
			items_by_code = {}
			for item in self.items:
				if not item.item_code:
					continue
				if item.item_code not in items_by_code:
					items_by_code[item.item_code] = {
						"item_name": item.item_name,
						"total_qty": flt(item.total_qty),
						"allocated_sum": 0.0
					}
				items_by_code[item.item_code]["allocated_sum"] += flt(item.allocated_qty)
			
			errors = []
			for code, details in items_by_code.items():
				if details["allocated_sum"] > details["total_qty"]:
					errors.append(
						_("For Item {0} ({1}), total allocated quota ({2}) cannot exceed Shipment Qty ({3}).")
						.format(code, details["item_name"], details["allocated_sum"], details["total_qty"])
					)
			
			if errors:
				# Show all errors in one go with clean layout
				error_msg = "<br>".join([f"&bull; {err}" for err in errors])
				frappe.throw(
					_("<b>Allocation Quota Violations:</b><br>{0}").format(error_msg),
					title=_("Validation Error")
				)

@frappe.whitelist()
def fetch_partner_requests(docname):
	doc = frappe.get_doc("Team Leader Allocation Sheet", docname)
	if doc.status == "Approved":
		frappe.throw(_("Cannot fetch requests for an Approved sheet."))
		
	doc.set("items", [])
	
	selected_shipments = [s.shipment_tracker for s in doc.shipments if s.shipment_tracker]
	if not selected_shipments:
		frappe.throw(_("Please select at least one Shipment Tracker first."))
		
	# Find all individual sheets in "Pending Team Leader"
	pending_sheets = frappe.get_all("Item Allocation Sheet", 
		filters={
			"status": "Pending Team Leader",
			"docstatus": 0
		}, 
		fields=["name", "sales_partner"])
		
	matched_count = 0
	for sheet_meta in pending_sheets:
		sheet = frappe.get_doc("Item Allocation Sheet", sheet_meta["name"])
		partner = sheet.sales_partner
		for item in sheet.items:
			if item.shipment in selected_shipments:
				child = doc.append("items", {})
				child.shipment = item.shipment
				child.item_code = item.item_code
				child.item_name = item.item_name
				child.description = item.description
				child.total_qty = flt(item.total_qty)
				child.sales_partner = partner
				child.customer = item.customer
				child.allocation_request = flt(item.allocation_request)
				# Initially allocate 0.0 (blank quota) as requested by user
				child.allocated_qty = 0.0
				child.source_doc = sheet.name
				child.source_row = item.name
				matched_count += 1
				
	doc.save()
	return f"Successfully fetched {matched_count} partner requests."

@frappe.whitelist()
def distribute_tl_quotas(docname):
	doc = frappe.get_doc("Team Leader Allocation Sheet", docname)
	if doc.status == "Approved":
		frappe.throw(_("Quotas have already been approved and distributed."))
		
	updated_docs = set()
	
	for item in doc.items:
		if item.source_doc and item.source_row:
			rows_list = item.source_row.split(",")
			# Set the allocated_qty of each individual row in the child table
			if len(rows_list) == 1:
				r_name = rows_list[0].strip()
				if r_name and frappe.db.exists("Partner Allocation Detail", r_name):
					frappe.db.set_value("Partner Allocation Detail", r_name, "allocated_qty", item.allocated_qty)
			else:
				# If there are multiple rows consolidated, distribute the allocated qty proportionally
				total_requested = flt(frappe.db.get_value("Partner Allocation Detail", {"name": ["in", rows_list]}, "sum(allocation_request)")) or 1.0
				remaining_allocated = flt(item.allocated_qty)
				for i, r_name in enumerate(rows_list):
					r_name = r_name.strip()
					if not r_name or not frappe.db.exists("Partner Allocation Detail", r_name):
						continue
					if i == len(rows_list) - 1:
						frappe.db.set_value("Partner Allocation Detail", r_name, "allocated_qty", remaining_allocated)
					else:
						req = flt(frappe.db.get_value("Partner Allocation Detail", r_name, "allocation_request"))
						share = flt((req / total_requested) * item.allocated_qty)
						frappe.db.set_value("Partner Allocation Detail", r_name, "allocated_qty", share)
						remaining_allocated -= share
						
			updated_docs.add(item.source_doc)
			
	for s_name in updated_docs:
		s_doc = frappe.get_doc("Item Allocation Sheet", s_name)
		s_doc.status = "Pending Partner Finalization"
		s_doc.save()
		
	doc.status = "Approved"
	doc.save()
	return "Quotas successfully approved and distributed to all Sales Partners."

@frappe.whitelist()
def get_item_spqs(docname, item_codes=None):
	if item_codes:
		if isinstance(item_codes, str):
			import json
			item_codes = json.loads(item_codes)
	else:
		if docname and not docname.startswith("new-team-leader-allocation-sheet-") and frappe.db.exists("Team Leader Allocation Sheet", docname):
			doc = frappe.get_doc("Team Leader Allocation Sheet", docname)
			item_codes = list(set([item.item_code for item in doc.items if item.item_code]))
		else:
			item_codes = []
	spqs = {}
	
	# Defensive check: does the custom field exist on Item?
	has_spq = frappe.get_meta("Item").has_field("custom_standard_packing_qty")
	
	if item_codes:
		if has_spq:
			items_data = frappe.get_all("Item", filters={"name": ["in", item_codes]}, fields=["name", "custom_standard_packing_qty"])
			for item in items_data:
				spqs[item.name] = item.custom_standard_packing_qty or 1
		else:
			for code in item_codes:
				spqs[code] = 1
				
	# Fetch all active Sales Partners defensively (avoiding OperationalError if sales_partner_name column is missing)
	meta = frappe.get_meta("Sales Partner")
	filters = {}
	if meta.has_field("disabled"):
		filters["disabled"] = 0
		
	fields = ["name"]
	if meta.has_field("sales_partner_name"):
		fields.append("sales_partner_name")
		
	partners = frappe.get_all("Sales Partner", filters=filters, order_by="name asc", fields=fields)
	
	return {
		"spqs": spqs,
		"partners": partners
	}

@frappe.whitelist()
def download_tl_template(docname=None):
	if not docname:
		frappe.throw(_("No document name provided."))
		
	doc = frappe.get_doc("Team Leader Allocation Sheet", docname)
	
	# Fetch all active Sales Partners defensively (avoiding OperationalError if sales_partner_name column is missing)
	meta = frappe.get_meta("Sales Partner")
	filters = {}
	if meta.has_field("disabled"):
		filters["disabled"] = 0
		
	fields = ["name"]
	has_sp_name = meta.has_field("sales_partner_name")
	if has_sp_name:
		fields.append("sales_partner_name")
		
	db_partners = frappe.get_all("Sales Partner", filters=filters, order_by="name asc", fields=fields)
	all_partners = [p.name for p in db_partners]
	
	partner_display_names = {}
	for p in db_partners:
		p_name = p.name
		if has_sp_name and p.get("sales_partner_name"):
			p_name = p.sales_partner_name
		partner_display_names[p.name] = p_name
	
	items_by_code = {}
	# Defensive check: does the custom field exist on Item?
	has_spq = frappe.get_meta("Item").has_field("custom_standard_packing_qty")
	
	for item in doc.items:
		if not item.item_code:
			continue
		key = item.item_code
		if key not in items_by_code:
			spq = 1
			if has_spq:
				spq = frappe.db.get_value("Item", item.item_code, "custom_standard_packing_qty") or 1
			items_by_code[key] = {
				"item_code": item.item_code,
				"item_name": item.item_name,
				"description": item.description,
				"total_qty": item.total_qty or 0,
				"spq": spq,
				"partners": {},
				"total_req": 0,
				"total_quota": 0
			}
		
		p = item.sales_partner
		if p not in items_by_code[key]["partners"]:
			items_by_code[key]["partners"][p] = {
				"request": 0,
				"quota": 0
			}
			
		partner_data = items_by_code[key]["partners"][p]
		partner_data["request"] += item.allocation_request or 0
		partner_data["quota"] += item.allocated_qty if item.allocated_qty is not None else 0
		
		items_by_code[key]["total_req"] += item.allocation_request or 0
		items_by_code[key]["total_quota"] += item.allocated_qty if item.allocated_qty is not None else 0

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "TL Quota Allocation"
	
	headers = ["Item Code", "Item Name", "Description", "Shipment Qty", "SPQ"]
	for p in all_partners:
		headers.append(partner_display_names[p])
	headers.extend(["Total Request", "Remaining Qty"])
	
	ws.append(headers)
	
	for key, details in items_by_code.items():
		row_data = [
			details["item_code"],
			details["item_name"],
			details["description"],
			details["total_qty"],
			details["spq"]
		]
		for p in all_partners:
			quota = details["partners"].get(p, {}).get("quota", 0)
			row_data.append(quota)
			
		remaining_qty = details["total_qty"] - details["total_quota"]
		row_data.append(details["total_req"])
		row_data.append(remaining_qty)
		ws.append(row_data)
		
	from openpyxl.styles import Font, PatternFill, Alignment
	
	header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
	header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
	center_align = Alignment(horizontal="center", vertical="center")
	left_align = Alignment(horizontal="left", vertical="center")
	
	for col_idx, cell in enumerate(ws[1], 1):
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = center_align if col_idx not in [2, 3] else left_align
		
	for col in ws.columns:
		vals = [cell.value for cell in col if cell.value is not None]
		max_len = max(len(str(v)) for v in vals) if vals else 10
		col_letter = openpyxl.utils.get_column_letter(col[0].column)
		ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
		
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	
	frappe.response["filename"] = f"TL_Quota_Allocation_{docname}.xlsx"
	frappe.response["filecontent"] = output.getvalue()
	frappe.response["type"] = "binary"

@frappe.whitelist()
def download_tl_request_template(docname=None):
	if not docname:
		frappe.throw(_("No document name provided."))
		
	doc = frappe.get_doc("Team Leader Allocation Sheet", docname)
	
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "TL Request Allocation"
	
	headers = ["Shipment", "Item Code", "Customer Name", "Item Name", "Description", "Sales Partner", "Allocation Request", "TL Quota"]
	ws.append(headers)
	
	# Check if Customer and Sales Partner have display names safely
	has_cust_name = False
	try:
		has_cust_name = frappe.get_meta("Customer").get_field("customer_name") is not None
	except Exception:
		pass
		
	has_sp_name = False
	try:
		has_sp_name = frappe.get_meta("Sales Partner").get_field("sales_partner_name") is not None
	except Exception:
		pass
		
	for item in doc.items:
		cust_id = item.customer
		cust_display = cust_id or ""
		if cust_id and has_cust_name:
			cust_display = frappe.db.get_value("Customer", cust_id, "customer_name") or cust_id
			
		partner_id = item.sales_partner
		partner_display = partner_id or ""
		if partner_id and has_sp_name:
			partner_display = frappe.db.get_value("Sales Partner", partner_id, "sales_partner_name") or partner_id
			
		row_data = [
			item.shipment or "",
			item.item_code or "",
			cust_display,
			item.item_name or "",
			item.description or "",
			partner_display,
			item.allocation_request or 0,
			item.allocated_qty if item.allocated_qty is not None else 0
		]
		ws.append(row_data)
		
	from openpyxl.styles import Font, PatternFill, Alignment
	
	header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
	header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
	center_align = Alignment(horizontal="center", vertical="center")
	left_align = Alignment(horizontal="left", vertical="center")
	
	for col_idx, cell in enumerate(ws[1], 1):
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = center_align if col_idx not in [3, 4, 5] else left_align
		
	for col in ws.columns:
		vals = [cell.value for cell in col if cell.value is not None]
		max_len = max(len(str(v)) for v in vals) if vals else 10
		col_letter = openpyxl.utils.get_column_letter(col[0].column)
		ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
		
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	
	frappe.response["filename"] = f"TL_Request_Allocation_{docname}.xlsx"
	frappe.response["filecontent"] = output.getvalue()
	frappe.response["type"] = "binary"

@frappe.whitelist()
def upload_tl_excel(docname):
	doc = frappe.get_doc("Team Leader Allocation Sheet", docname)
	if doc.status == "Approved":
		frappe.throw(_("Cannot upload Excel data to an Approved sheet."))
		
	if not doc.excel_file:
		frappe.throw(_("Please upload an Excel file first."))
		
	file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
	file_path = file_doc.get_full_path()
	
	wb = openpyxl.load_workbook(file_path, data_only=True)
	ws = wb.active
	
	rows = list(ws.iter_rows(values_only=True))
	if not rows:
		frappe.throw(_("Excel sheet is empty."))
		
	raw_headers = [str(h).strip() for h in rows[0]]
	headers_lower = [h.lower() for h in raw_headers]
	
	is_request_template = "sales partner" in headers_lower and ("tl quota" in headers_lower or "allocated qty" in headers_lower)
	
	if is_request_template:
		if "item code" not in headers_lower:
			frappe.throw(_("Excel template is missing required column: Item Code"))
			
		item_code_idx = headers_lower.index("item code")
		partner_idx = headers_lower.index("sales partner")
		
		customer_idx = None
		if "customer name" in headers_lower:
			customer_idx = headers_lower.index("customer name")
		elif "customer" in headers_lower:
			customer_idx = headers_lower.index("customer")
			
		quota_idx = None
		if "tl quota" in headers_lower:
			quota_idx = headers_lower.index("tl quota")
		elif "allocated qty" in headers_lower:
			quota_idx = headers_lower.index("allocated qty")
			
		if quota_idx is None:
			frappe.throw(_("Excel template is missing required column: TL Quota"))
			
		updated_count = 0
		for row in rows[1:]:
			item_code = str(row[item_code_idx] or "").strip()
			if not item_code:
				continue
				
			partner_name = str(row[partner_idx] or "").strip()
			if not partner_name:
				continue
				
			row_cust = ""
			if customer_idx is not None and customer_idx < len(row):
				row_cust = str(row[customer_idx] or "").strip()
				
			val = flt(row[quota_idx])
			
			# Find all matching child rows
			matching_children = []
			for child in doc.items:
				if (normalize_str(child.item_code) == normalize_str(item_code) and
					partners_match(child.sales_partner, partner_name) and
					customers_match(child.customer, row_cust)):
					matching_children.append(child)
					
			if len(matching_children) == 1:
				matching_children[0].allocated_qty = val
				updated_count += 1
			elif len(matching_children) > 1:
				total_requested = sum(flt(c.allocation_request) for c in matching_children) or 1.0
				remaining_val = val
				for idx, child in enumerate(matching_children):
					if idx == len(matching_children) - 1:
						child.allocated_qty = remaining_val
					else:
						share = flt((flt(child.allocation_request) / total_requested) * val)
						child.allocated_qty = share
						remaining_val -= share
					updated_count += 1
					
		doc.save()
		return f"Successfully updated {updated_count} partner quota entries from Request Excel."
		
	else:
		if "item code" not in headers_lower:
			frappe.throw(_("Excel template is missing required column: Item Code"))
				
		item_code_idx = headers_lower.index("item code")
		
		static_cols = ["item code", "item name", "description", "shipment qty", "spq", "total request", "remaining qty"]
		
		partner_cols = []
		for idx, h in enumerate(raw_headers):
			h_lower = h.lower()
			if h_lower not in static_cols:
				partner_cols.append({
					"partner_name": h,
					"index": idx
				})
				
		if not partner_cols:
			frappe.throw(_("Excel sheet does not contain any Sales Partner columns."))
			
		updated_count = 0
		for row in rows[1:]:
			item_code = str(row[item_code_idx] or "").strip()
			if not item_code:
				continue
				
			for p_col in partner_cols:
				partner_name = p_col["partner_name"]
				val_idx = p_col["index"]
				
				val = 0
				if val_idx < len(row):
					val = flt(row[val_idx])
					
				# Find all matching child rows purely by item_code and sales_partner
				matching_children = []
				for child in doc.items:
					if (normalize_str(child.item_code) == normalize_str(item_code) and
						partners_match(child.sales_partner, partner_name)):
						matching_children.append(child)
						
				if len(matching_children) == 1:
					matching_children[0].allocated_qty = val
					updated_count += 1
				elif len(matching_children) > 1:
					total_requested = sum(flt(c.allocation_request) for c in matching_children) or 1.0
					remaining_val = val
					for idx, child in enumerate(matching_children):
						if idx == len(matching_children) - 1:
							child.allocated_qty = remaining_val
						else:
							share = flt((flt(child.allocation_request) / total_requested) * val)
							child.allocated_qty = share
							remaining_val -= share
						updated_count += 1
				elif val > 0:
					# No existing child rows for this partner and item! Create a new row!
					child = doc.append("items", {})
					child.item_code = item_code
					
					# Find the sales partner ID matching this partner_name
					partner_id = partner_name
					has_sp_name = False
					try:
						has_sp_name = frappe.get_meta("Sales Partner").get_field("sales_partner_name") is not None
					except Exception:
						pass
					if has_sp_name:
						res = frappe.db.get_value("Sales Partner", {"sales_partner_name": partner_name}, "name")
						if res:
							partner_id = res
							
					child.sales_partner = partner_id
					child.allocated_qty = val
					child.allocation_request = 0.0
					
					# Get item details from existing rows in doc.items
					match_item = next((i for i in doc.items if i.item_code == item_code and i.item_name), None)
					if match_item:
						child.item_name = match_item.item_name
						child.description = match_item.description
						child.total_qty = match_item.total_qty
						child.shipment = match_item.shipment
					updated_count += 1
						
		doc.save()
		return f"Successfully updated partner quota entries from Matrix Excel."